import re
import pandas as pd

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# COLUMN HELPERS
# =============================================================================

def normalize_columns(df):
    return df.rename(columns={c: str(c).strip().lower() for c in df.columns})

def wave_code(val):
    """Accept '22', 22, 'w22', 'W22' — always return 'w22' format."""
    if pd.isna(val):
        return ""
    s = str(val).strip().lower()
    if s.startswith("w"):
        try:
            int(s[1:])
            return "w" + s[1:]
        except ValueError:
            return s
    try:
        return f"w{int(float(s))}"
    except ValueError:
        return s

def strip_parcel_suffix(code):
    if pd.isna(code):
        return ""
    return re.sub(r'_p\d+$', '', str(code))

def make_geometry(row):
    lat = row.get("latitude")
    lon = row.get("longitude")
    if pd.notna(lat) and pd.notna(lon):
        return f"{lat} {lon} 0.0 0.0"
    return ""


# =============================================================================
# EXCEL STYLING
# =============================================================================

def _header_style(cell, bg="4F81BD"):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def style_sheet(ws, df):
    for cell in ws[1]:
        _header_style(cell)
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

def to_excel(dfs: dict) -> bytes:
    """Write multiple DataFrames into a styled multi-sheet Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(sheet_name[:31])   # Excel sheet name limit
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        style_sheet(ws, df)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def enforce_label_fallback(df):
    """If 'label' column exists, fill empty values from 'name' column."""
    if "label" not in df.columns:
        return df
    df = df.copy()
    if "name" in df.columns:
        mask = df["label"].isna() | df["label"].astype(str).str.strip().eq("")
        df.loc[mask, "label"] = df.loc[mask, "name"]
    return df