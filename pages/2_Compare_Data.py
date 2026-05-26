import streamlit as st
import pandas as pd

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Compare Data",
    page_icon="🔍",
    layout="wide"
)

st.title("🔍 Compare Data")
st.caption(
    "Upload two Excel files, pick a column from each, "
    "and see how many values from File A exist in File B — and what's missing."
)

# =============================================================================
# HELPERS
# =============================================================================

def load_sheets(file):
    xl = pd.ExcelFile(file)
    return {name: xl.parse(name) for name in xl.sheet_names}

def style_ws(ws, df, header_color="4F81BD"):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

def build_report_excel(summary_df, matched_df, missing_df, multi_match_df) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Summary",       summary_df,     "1F4E79"),
        ("Matched",       matched_df,     "375623"),
        ("Missing",       missing_df,     "843C0C"),
        ("Multi_Match",   multi_match_df, "7030A0"),
    ]

    for sheet_name, df, color in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        style_ws(ws, df, color)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =============================================================================
# UPLOAD
# =============================================================================

col_a, col_b = st.columns(2)

with col_a:
    st.markdown("### File A (source)")
    file_a = st.file_uploader("Upload File A (.xlsx)", type=["xlsx"], key="file_a")

with col_b:
    st.markdown("### File B (reference)")
    file_b = st.file_uploader("Upload File B (.xlsx)", type=["xlsx"], key="file_b")

if not file_a or not file_b:
    st.info("⬆️ Upload both files to continue.")
    st.stop()

# =============================================================================
# LOAD & SHEET SELECTION
# =============================================================================

@st.cache_data(show_spinner=False)
def cached_load(file):
    return load_sheets(file)

with st.spinner("Reading files..."):
    sheets_a = cached_load(file_a)
    sheets_b = cached_load(file_b)

st.success(f"File A: **{len(sheets_a)} sheet(s)** | File B: **{len(sheets_b)} sheet(s)**")

col1, col2 = st.columns(2)

with col1:
    sheet_a_name = st.selectbox("Sheet from File A", list(sheets_a.keys()), key="sheet_a")
    df_a = sheets_a[sheet_a_name]
    st.caption(f"{len(df_a):,} rows × {len(df_a.columns)} columns")

with col2:
    sheet_b_name = st.selectbox("Sheet from File B", list(sheets_b.keys()), key="sheet_b")
    df_b = sheets_b[sheet_b_name]
    st.caption(f"{len(df_b):,} rows × {len(df_b.columns)} columns")

# =============================================================================
# COLUMN SELECTION — multiple columns allowed
# =============================================================================

st.subheader("Column Selection")
st.caption(
    "Select one or more columns to compare. "
    "If multiple columns are selected, values are combined into a single key for matching."
)

col3, col4 = st.columns(2)

with col3:
    cols_a = st.multiselect(
        "Column(s) from File A",
        options=list(df_a.columns),
        key="cols_a"
    )

with col4:
    cols_b = st.multiselect(
        "Column(s) from File B",
        options=list(df_b.columns),
        key="cols_b"
    )

if not cols_a or not cols_b:
    st.info("Select at least one column from each file.")
    st.stop()

# =============================================================================
# RUN COMPARISON
# =============================================================================

if st.button("🔍 Run Comparison", type="primary"):

    # Build composite keys (concat multiple cols with | separator)
    def make_key(df, cols):
        return df[cols].fillna("").astype(str).agg(" | ".join, axis=1).str.strip()

    key_a = make_key(df_a, cols_a)
    key_b = make_key(df_b, cols_b)

    # Unique values and sets
    unique_a = key_a.dropna().unique()
    set_b    = set(key_b.dropna().unique())

    matched_vals = [v for v in unique_a if v in set_b]
    missing_vals = [v for v in unique_a if v not in set_b]

    n_unique_a  = len(unique_a)
    n_matched   = len(matched_vals)
    n_missing   = len(missing_vals)
    pct_matched = round(n_matched / n_unique_a * 100, 1) if n_unique_a else 0
    pct_missing = round(n_missing / n_unique_a * 100, 1) if n_unique_a else 0

    # Count occurrences in A for matched and missing
    counts_a = key_a.value_counts().rename("count_in_A")

    # ── Summary ───────────────────────────────────────────────────────────────
    summary_df = pd.DataFrame([
        {"Metric": "File A — sheet",               "Value": sheet_a_name},
        {"Metric": "File B — sheet",               "Value": sheet_b_name},
        {"Metric": "Column(s) from A",             "Value": " | ".join(cols_a)},
        {"Metric": "Column(s) from B",             "Value": " | ".join(cols_b)},
        {"Metric": "Total rows in A",              "Value": len(df_a)},
        {"Metric": "Total rows in B",              "Value": len(df_b)},
        {"Metric": "Unique values in A",           "Value": n_unique_a},
        {"Metric": "Unique values in B",           "Value": len(set_b)},
        {"Metric": "Matched (A exists in B)",      "Value": f"{n_matched} ({pct_matched}%)"},
        {"Metric": "Missing (A not found in B)",   "Value": f"{n_missing} ({pct_missing}%)"},
    ])

    # ── Matched detail ────────────────────────────────────────────────────────
    matched_df = pd.DataFrame({
        "key":        matched_vals,
        "count_in_A": [int(counts_a.get(v, 0)) for v in matched_vals],
    }).sort_values("count_in_A", ascending=False).reset_index(drop=True)
    matched_df.columns = ["value", "count_in_A"]

    # ── Missing detail ────────────────────────────────────────────────────────
    missing_df = pd.DataFrame({
        "key":        missing_vals,
        "count_in_A": [int(counts_a.get(v, 0)) for v in missing_vals],
    }).sort_values("count_in_A", ascending=False).reset_index(drop=True)
    missing_df.columns = ["value", "count_in_A"]

    # ── Multi-match: values in A that appear more than once ───────────────────
    dupes_a = counts_a[counts_a > 1].reset_index()
    dupes_a.columns = ["value", "count_in_A"]
    dupes_a["found_in_B"] = dupes_a["value"].apply(lambda v: "Yes" if v in set_b else "No")
    multi_match_df = dupes_a.sort_values("count_in_A", ascending=False).reset_index(drop=True)

    # Store in session
    st.session_state["comparison"] = {
        "summary_df":     summary_df,
        "matched_df":     matched_df,
        "missing_df":     missing_df,
        "multi_match_df": multi_match_df,
        "n_matched":      n_matched,
        "n_missing":      n_missing,
        "n_unique_a":     n_unique_a,
        "pct_matched":    pct_matched,
        "pct_missing":    pct_missing,
    }

# =============================================================================
# RESULTS
# =============================================================================

if "comparison" not in st.session_state:
    st.stop()

res = st.session_state["comparison"]

st.subheader("Results")

m1, m2, m3, m4 = st.columns(4)
m1.metric("Unique values in A",  f"{res['n_unique_a']:,}")
m2.metric("Matched in B",        f"{res['n_matched']:,}",  delta=f"{res['pct_matched']}%")
m3.metric("Missing from B",      f"{res['n_missing']:,}",  delta=f"-{res['pct_missing']}%", delta_color="inverse")
m4.metric("Match rate",          f"{res['pct_matched']}%")

st.markdown("")

tab_sum, tab_match, tab_miss, tab_multi = st.tabs([
    "📋 Summary",
    f"✅ Matched ({res['n_matched']:,})",
    f"❌ Missing ({res['n_missing']:,})",
    "🔢 Multi-occurrence in A",
])

with tab_sum:
    st.dataframe(res["summary_df"], use_container_width=True, hide_index=True)

with tab_match:
    st.caption("Values from File A that were found in File B, with their occurrence count in A.")
    st.dataframe(res["matched_df"], use_container_width=True, hide_index=True)

with tab_miss:
    st.caption("Values from File A that were NOT found in File B.")
    if res["missing_df"].empty:
        st.success("🎉 No missing values — all values from A exist in B!")
    else:
        st.dataframe(res["missing_df"], use_container_width=True, hide_index=True)

with tab_multi:
    st.caption("Values that appear more than once in File A (duplicates), and whether they exist in B.")
    if res["multi_match_df"].empty:
        st.info("No duplicate values in File A for the selected column(s).")
    else:
        st.dataframe(res["multi_match_df"], use_container_width=True, hide_index=True)

# =============================================================================
# DOWNLOAD REPORT
# =============================================================================

st.subheader("Download Report")

report_bytes = build_report_excel(
    res["summary_df"],
    res["matched_df"],
    res["missing_df"],
    res["multi_match_df"],
)

st.download_button(
    label="⬇️ Download Comparison Report (.xlsx)",
    data=report_bytes,
    file_name="comparison_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)