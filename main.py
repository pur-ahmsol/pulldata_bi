import re
import os
import json
import zipfile
import tempfile

import streamlit as st
import pandas as pd
import pydeck as pdk

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# OPTIONAL GIS LIBRARIES
# =============================================================================

try:
    import geopandas as gpd
    from shapely.geometry import Point, Polygon
    GIS_AVAILABLE = True
except Exception:
    GIS_AVAILABLE = False

# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="Pulldata Biomass Tools",
    page_icon="🪵",
    layout="wide"
)

st.title("🪵 Pulldata Biomass Tools")
st.caption("Transform RAW survey data into structured pulldata tabs and GIS exports")

# =============================================================================
# REQUIRED VARIABLES
# =============================================================================

REQUIRED_COLUMNS = [
    ("farmerid",          "Farmer ID"),
    ("farmer_name",       "Farmer Name"),
    ("code",              "Parcel Code"),
    ("subproject_name",   "Subproject"),
    ("location_level_3",  "Location"),
    ("project_id",        "Project ID"),
    ("wave_u",            "Wave"),
    ("id",                "Parcel ID"),
    ("new_id",            "New ID"),
    ("cr_parcel_landuse", "Strata"),
    ("plantmodel",        "Planting Model"),
    ("point_type",        "Point Type"),
    ("latitude",          "Latitude"),
    ("longitude",         "Longitude"),
]

# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.header("🪵 Biomass Tools")
    st.markdown("""
### 🏠 Pages
- **Pulldata Biomass Tools** ← you are here
- **Extract CSV** — export Excel sheets to CSV
- **Compare Data** — match values across files

### Outputs
- farmer_name
- parcel
- pulldata_sampling_biomass
- subplot_gpspoint
- GIS export
""")
    if GIS_AVAILABLE:
        st.success("GIS libraries loaded")
    else:
        st.warning("GeoPandas unavailable")

# =============================================================================
# HELPERS
# =============================================================================

def normalize_columns(df):
    return df.rename(columns={c: str(c).strip().lower() for c in df.columns})

def validate_raw_columns(df):
    existing = set(df.columns)
    return [(col, desc) for col, desc in REQUIRED_COLUMNS if col not in existing]

def wave_code(val):
    """Accept '22', 22, 'w22', 'W22' — always return 'w22' format."""
    if pd.isna(val):
        return ""
    s = str(val).strip().lower()
    if s.startswith("w"):
        # already formatted, normalise casing
        try:
            int(s[1:])   # validate the numeric part
            return "w" + s[1:]
        except ValueError:
            return s
    try:
        return f"w{int(float(s))}"
    except ValueError:
        return s

def strip_parcel_suffix(code):
    if pd.isna(code):
        return ""
    return re.sub(r'_p\d+$', '', str(code))

def make_geometry(row):
    lat = row.get("latitude")
    lon = row.get("longitude")
    if pd.notna(lat) and pd.notna(lon):
        return f"{lat} {lon} 0.0 0.0"
    return ""

def build_subproject_lookup(raw):
    if "subproject_id" not in raw.columns:
        return {}
    grouped = (
        raw[["subproject_name", "subproject_id"]]
        .dropna(subset=["subproject_name", "subproject_id"])
        .groupby("subproject_name")["subproject_id"]
        .first()
    )
    return grouped.to_dict()

def detect_project_name(raw):
    """Return the first non-null project_name from RAW, or empty string."""
    for col in ["project_name", "projectname"]:
        if col in raw.columns:
            vals = raw[col].dropna().astype(str).str.strip()
            vals = vals[vals != ""]
            if not vals.empty:
                return vals.iloc[0]
    return ""

# =============================================================================
# TRACK PARSER
# =============================================================================

def parse_track_geometry(geometry_text):
    try:
        coords = []
        segments = str(geometry_text).split(";")
        for seg in segments:
            seg = seg.strip()
            if not seg:
                continue
            parts = seg.split()
            lat = float(parts[0])
            lon = float(parts[1])
            coords.append([lon, lat])
        return coords
    except:
        return None

# =============================================================================
# MAP BUILDERS
# =============================================================================

def build_point_layer(raw):
    rows = []
    valid = raw[["latitude", "longitude"]].dropna()
    for _, row in valid.iterrows():
        try:
            rows.append({"lat": float(row["latitude"]), "lon": float(row["longitude"])})
        except:
            pass
    return pd.DataFrame(rows)

def build_polygon_layer(track_df):
    polygons = []
    if track_df is None:
        return pd.DataFrame()
    for _, row in track_df.iterrows():
        coords = parse_track_geometry(row.get("geometry"))
        if not coords or len(coords) < 4:
            continue
        if coords[0] != coords[-1]:
            coords.append(coords[0])
        polygons.append({"polygon": coords, "label": row.get("label", "")})
    return pd.DataFrame(polygons)

# =============================================================================
# GEOJSON / KML / SHP
# =============================================================================

def build_combined_geojson(raw, track_df=None):
    features = []

    def sv(row, key):
        val = row.get(key, "")
        return "" if pd.isna(val) else str(val).strip()

    valid = raw[["latitude", "longitude"]].dropna()
    for _, row in valid.iterrows():
        try:
            features.append({
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [float(row["longitude"]), float(row["latitude"])]
                },
                "properties": {
                    "layer":           "subplot_gpspoint",
                    "farmer_name":     sv(row, "farmer_name"),
                    "wave":            wave_code(row.get("wave_u")),
                    "subproject_name": sv(row, "subproject_name"),
                    "parcel_code":     sv(row, "code"),
                    "new_id":          sv(row, "new_id"),
                }
            })
        except:
            pass

    if track_df is not None:
        # Build lookup from RAW keyed on parcel code
        attr_cols = [c for c in ["code", "farmer_name", "wave_u", "subproject_name", "new_id"] if c in raw.columns]
        raw_lookup = raw[attr_cols].drop_duplicates(subset=["code"] if "code" in attr_cols else attr_cols[:1])

        for _, row in track_df.iterrows():
            coords = parse_track_geometry(row.get("geometry"))
            if not coords or len(coords) < 4:
                continue
            if coords[0] != coords[-1]:
                coords.append(coords[0])

            label = str(row.get("label", "") or row.get("parcel_code", "")).strip()
            props = {"layer": "parcel_polygon", "label": label}

            if not raw_lookup.empty and label and "code" in raw_lookup.columns:
                match = raw_lookup[raw_lookup["code"].astype(str).str.strip() == label]
                if not match.empty:
                    mr = match.iloc[0]
                    props["farmer_name"]     = sv(mr, "farmer_name")
                    props["wave"]            = wave_code(mr.get("wave_u"))
                    props["subproject_name"] = sv(mr, "subproject_name")
                    props["parcel_code"]     = sv(mr, "code")
                    props["new_id"]          = sv(mr, "new_id")

            features.append({
                "type": "Feature",
                "geometry": {"type": "Polygon", "coordinates": [coords]},
                "properties": props
            })

    return json.dumps({"type": "FeatureCollection", "features": features}, indent=2).encode("utf-8")

def _kml_extended_data(props, skip=("layer",)):
    """Render a KML <ExtendedData> block from a dict of properties."""
    lines = ["<ExtendedData>"]
    for k, v in props.items():
        if k in skip:
            continue
        safe_v = str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        lines.append(f'  <Data name="{k}"><value>{safe_v}</value></Data>')
    lines.append("</ExtendedData>")
    return lines

def geojson_to_kml(geojson_bytes):
    geojson = json.loads(geojson_bytes.decode("utf-8"))
    kml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        "<Document>",
    ]
    for feature in geojson["features"]:
        geom  = feature["geometry"]
        props = feature.get("properties", {})
        name  = props.get("new_id") or props.get("label") or ""
        if geom["type"] == "Point":
            lon, lat = geom["coordinates"]
            kml += ["<Placemark>", f"<name>{name}</name>"]
            kml += _kml_extended_data(props)
            kml += ["<Point>", f"<coordinates>{lon},{lat},0</coordinates>", "</Point>", "</Placemark>"]
        elif geom["type"] == "Polygon":
            coords_text = " ".join([f"{c[0]},{c[1]},0" for c in geom["coordinates"][0]])
            kml += ["<Placemark>", f"<name>{name}</name>"]
            kml += _kml_extended_data(props)
            kml += ["<Polygon>", "<outerBoundaryIs>", "<LinearRing>",
                    f"<coordinates>{coords_text}</coordinates>",
                    "</LinearRing>", "</outerBoundaryIs>", "</Polygon>", "</Placemark>"]
    kml += ["</Document>", "</kml>"]
    return "\n".join(kml).encode("utf-8")

def _sv(row, key):
    val = row.get(key, "")
    return "" if pd.isna(val) else str(val).strip()

def export_shapefile(raw, track_df):
    if not GIS_AVAILABLE:
        return None
    temp_dir = tempfile.mkdtemp()

    # ── Points ────────────────────────────────────────────────────────────────
    point_records = []
    valid = raw[["latitude", "longitude"]].dropna()
    for _, row in valid.iterrows():
        try:
            point_records.append({
                "geometry":        Point(float(row["longitude"]), float(row["latitude"])),
                "layer":           "subplot_gpspoint",
                "farmer_name":     _sv(row, "farmer_name"),
                "wave":            wave_code(row.get("wave_u")),
                "subproject":      _sv(row, "subproject_name"),
                "parcel_code":     _sv(row, "code"),
                "new_id":          _sv(row, "new_id"),
            })
        except:
            pass
    if point_records:
        pts_df = pd.DataFrame(point_records)
        gpd.GeoDataFrame(pts_df, geometry="geometry", crs="EPSG:4326").to_file(
            os.path.join(temp_dir, "subplot_points.shp")
        )

    # ── Polygons ──────────────────────────────────────────────────────────────
    # build RAW lookup by parcel code
    attr_cols = [c for c in ["code", "farmer_name", "wave_u", "subproject_name", "new_id"] if c in raw.columns]
    raw_lookup = raw[attr_cols].drop_duplicates(subset=["code"] if "code" in attr_cols else attr_cols[:1]) if attr_cols else pd.DataFrame()

    polygon_records = []
    if track_df is not None:
        for _, row in track_df.iterrows():
            coords = parse_track_geometry(row.get("geometry"))
            if not coords or len(coords) < 4:
                continue
            try:
                if coords[0] != coords[-1]:
                    coords.append(coords[0])
                label = str(row.get("label", "") or row.get("parcel_code", "")).strip()
                rec = {
                    "geometry":    Polygon(coords),
                    "layer":       "parcel_polygon",
                    "label":       label,
                    "farmer_name": "",
                    "wave":        "",
                    "subproject":  "",
                    "parcel_code": label,
                    "new_id":      "",
                }
                if not raw_lookup.empty and label and "code" in raw_lookup.columns:
                    match = raw_lookup[raw_lookup["code"].astype(str).str.strip() == label]
                    if not match.empty:
                        mr = match.iloc[0]
                        rec["farmer_name"] = _sv(mr, "farmer_name")
                        rec["wave"]        = wave_code(mr.get("wave_u"))
                        rec["subproject"]  = _sv(mr, "subproject_name")
                        rec["new_id"]      = _sv(mr, "new_id")
                polygon_records.append(rec)
            except:
                pass
    if polygon_records:
        poly_df = pd.DataFrame(polygon_records)
        gpd.GeoDataFrame(poly_df, geometry="geometry", crs="EPSG:4326").to_file(
            os.path.join(temp_dir, "parcel_polygons.shp")
        )

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in os.listdir(temp_dir):
            zf.write(os.path.join(temp_dir, file), arcname=file)
    zip_buffer.seek(0)
    return zip_buffer

# =============================================================================
# EXCEL EXPORT
# =============================================================================

def style_sheet(ws, df):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

def enforce_label_fallback(df):
    """If 'label' column exists, fill any empty/null values from 'name' column."""
    if "label" not in df.columns:
        return df
    df = df.copy()
    if "name" in df.columns:
        mask = df["label"].isna() | df["label"].astype(str).str.strip().eq("")
        df.loc[mask, "label"] = df.loc[mask, "name"]
    return df

def to_excel(dfs):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in dfs.items():
        df = enforce_label_fallback(df)
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        style_sheet(ws, df)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =============================================================================
# TRANSFORMATION FUNCTIONS
# =============================================================================

def build_farmer_name(raw, subproject_lookup):
    seen = set()
    rows = []
    for _, r in raw.iterrows():
        fid = r.get("farmerid")
        if fid in seen or pd.isna(fid):
            continue
        seen.add(fid)
        sname = str(r.get("subproject_name", "")).strip()
        _farmer_label = str(r.get("farmer_name", "")).strip() or strip_parcel_suffix(r.get("code"))
        rows.append({
            "list_name":     "farmer",
            "name":          fid,
            "label":         _farmer_label,
            "organization":  sname,
            "location":      r.get("location_level_3", ""),
            "projectid":     r.get("project_id", ""),
            "subprojectid":  subproject_lookup.get(sname, ""),
            "producer_code": strip_parcel_suffix(r.get("code")),
            "farmer_name":   r.get("farmer_name", ""),
            "wave_code":     wave_code(r.get("wave_u")),
        })
    return pd.DataFrame(rows)

def build_parcel(raw):
    seen = set()
    rows = []
    for _, r in raw.iterrows():
        pid = r.get("id")
        if pid in seen or pd.isna(pid):
            continue
        seen.add(pid)
        rows.append({
            "list_name":    "parcel",
            "name":         pid,
            "label":        r.get("code", ""),
            "farmer":       r.get("farmerid", ""),
            "filter":       pid,
            "farmer_code":  strip_parcel_suffix(r.get("code")),
            "parcel_label": r.get("code", ""),
        })
    return pd.DataFrame(rows)

def build_pulldata(raw, subproject_lookup, project_name=""):
    rows = []
    for _, r in raw.iterrows():
        sname = str(r.get("subproject_name", "")).strip()
        rows.append({
            "project_name":    project_name,
            "project_id":      r.get("project_id", ""),
            "subproject_id":   subproject_lookup.get(sname, ""),
            "subproject_name": sname,
            "wave_code":       wave_code(r.get("wave_u")),
            "id_parcel":       r.get("id", ""),
            "parcel_code":     r.get("new_id", ""),
            "id_farmer":       r.get("farmerid", ""),
            "farmer_code":     strip_parcel_suffix(r.get("code")),
            "name":            r.get("new_id", ""),
            "subplot_yn":      "yes",
            "strata":          r.get("cr_parcel_landuse", ""),
            "planting_model":  r.get("plantmodel", ""),
            "shape":           "circular",
            "geometry":        make_geometry(r),
        })
    return pd.DataFrame(rows)

def build_subplot_gpspoint(raw):
    rows = []
    for _, r in raw.iterrows():
        rows.append({
            "name":      r.get("new_id", ""),
            "label":     r.get("new_id", ""),
            "wave_code": wave_code(r.get("wave_u")),
            "parcel_code": r.get("code", ""),
            "id_parcel": r.get("id", ""),
            "geometry":  make_geometry(r),
        })
    return pd.DataFrame(rows)

def build_subplot_and_parcel(raw, track_df):
    """
    Combined GPS export tab: all RAW subplot GPS points + parcel polygon
    track points, in the same column format as subplot_gpspoint.
    """
    rows = []

    # ── Part 1: RAW subplot GPS points ───────────────────────────────────────
    for _, r in raw.iterrows():
        rows.append({
            "source":                  "subplot_gpspoint",
            "name":                    r.get("new_id", ""),
            "label":                   r.get("new_id", ""),
            "wave_code":               wave_code(r.get("wave_u")),
            "parcel_code":             r.get("code", ""),
            "farmer_name":             r.get("farmer_name", ""),
            "subproject_name":         r.get("subproject_name", ""),
            "id_parcel":               r.get("id", ""),
            "geometry":                make_geometry(r),
        })

    # ── Part 2: parcel GPS track points ──────────────────────────────────────
    if track_df is not None:
        # build RAW lookup by parcel code for attribute enrichment
        attr_cols = [c for c in ["code", "farmer_name", "wave_u", "subproject_name", "id"] if c in raw.columns]
        raw_lookup = (
            raw[attr_cols].drop_duplicates(subset=["code"] if "code" in attr_cols else attr_cols[:1])
            if attr_cols else pd.DataFrame()
        )

        for _, r in track_df.iterrows():
            label = str(r.get("label", "") or r.get("parcel_code", "")).strip()
            geometry_str = str(r.get("geometry", "")).strip()

            # defaults
            farmer_name = ""
            wc = ""
            subproject_name = ""
            id_parcel = ""

            if not raw_lookup.empty and label and "code" in raw_lookup.columns:
                match = raw_lookup[raw_lookup["code"].astype(str).str.strip() == label]
                if not match.empty:
                    mr = match.iloc[0]
                    farmer_name     = str(mr.get("farmer_name", "") or "").strip()
                    wc              = wave_code(mr.get("wave_u"))
                    subproject_name = str(mr.get("subproject_name", "") or "").strip()
                    id_parcel       = mr.get("id", "")

            rows.append({
                "source":                  "parcel_gps_track",
                "name":                    label,
                "label":                   label,
                "wave_code":               wc,
                "parcel_code":             label,
                "farmer_name":             farmer_name,
                "subproject_name":         subproject_name,
                "id_parcel":               id_parcel,
                "geometry":                geometry_str,
            })

    return pd.DataFrame(rows)

# =============================================================================
# EMPTY COLUMN CHECKER
# =============================================================================

def check_empty_columns(results, raw):
    whole_empty = []
    partial_empty = []
    raw_columns = set(raw.columns)
    for sheet_name, df in results.items():
        for col in df.columns:
            missing_count = df[col].isna().sum()
            empty_string_count = df[col].astype(str).str.strip().eq("").sum()
            total_missing = max(missing_count, empty_string_count)
            if total_missing == len(df):
                whole_empty.append({"tab": sheet_name, "column": col})
            elif total_missing > 0:
                possible_raw = col if col in raw_columns else "Check source RAW variable"
                partial_empty.append({
                    "tab": sheet_name,
                    "column": col,
                    "missing_rows": total_missing,
                    "check_raw_variable": possible_raw
                })
    return whole_empty, partial_empty

# =============================================================================
# RAW UPLOAD
# =============================================================================

uploaded = st.file_uploader("Upload RAW Excel (.xlsx)", type=["xlsx"])

if not uploaded:
    st.stop()

@st.cache_data(show_spinner=False)
def load_raw(file):
    df = pd.read_excel(file, sheet_name="RAW")
    return normalize_columns(df)

raw = load_raw(uploaded)

# =============================================================================
# VALIDATION
# =============================================================================

st.subheader("Validation")

missing_cols = validate_raw_columns(raw)
manual_mapping = {}

if missing_cols:
    st.warning("Some required variables are missing.")
    existing_cols = sorted(raw.columns.tolist())
    for missing_col, desc in missing_cols:
        selected = st.selectbox(
            f"{missing_col} ({desc})",
            options=[""] + existing_cols,
            key=f"map_{missing_col}"
        )
        if selected:
            manual_mapping[missing_col] = selected
    unresolved = [c for c, _ in missing_cols if c not in manual_mapping]
    if unresolved:
        st.stop()
    for target_col, source_col in manual_mapping.items():
        raw[target_col] = raw[source_col]
    st.success("Mappings applied.")
else:
    st.success("All required variables found")

# =============================================================================
# DATA SUMMARY
# =============================================================================

st.subheader("Data Summary")

n_farmers  = raw["farmerid"].dropna().nunique()
n_parcels  = raw["id"].dropna().nunique()
n_points   = raw["new_id"].dropna().nunique()
n_subproj  = raw["subproject_name"].dropna().nunique() if "subproject_name" in raw.columns else 0
n_waves    = raw["wave_u"].dropna().nunique() if "wave_u" in raw.columns else 0
n_rows     = len(raw)

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Total Rows",       f"{n_rows:,}")
c2.metric("Unique Farmers",   f"{n_farmers:,}")
c3.metric("Unique Parcels",   f"{n_parcels:,}")
c4.metric("GPS Points",       f"{n_points:,}")
c5.metric("Subprojects",      f"{n_subproj:,}")
c6.metric("Waves",            f"{n_waves:,}")

# Subproject breakdown table
if "subproject_name" in raw.columns:
    with st.expander("Subproject breakdown"):
        breakdown = (
            raw.groupby("subproject_name")
            .agg(
                farmers=("farmerid", "nunique"),
                parcels=("id", "nunique"),
                gps_points=("new_id", "nunique"),
            )
            .reset_index()
            .rename(columns={"subproject_name": "Subproject"})
        )
        # attach subproject_id if available
        if "subproject_id" in raw.columns:
            sid_map = (
                raw[["subproject_name", "subproject_id"]]
                .dropna(subset=["subproject_id"])
                .groupby("subproject_name")["subproject_id"]
                .first()
                .reset_index()
                .rename(columns={"subproject_name": "Subproject", "subproject_id": "subproject_id"})
            )
            breakdown = breakdown.merge(sid_map, on="Subproject", how="left")
        st.dataframe(breakdown, use_container_width=True, hide_index=True)

# =============================================================================
# SUBPROJECT ID MAPPING
# =============================================================================

st.subheader("Subproject ID Mapping")

_auto_lookup = build_subproject_lookup(raw)
_unique_subprojects = (
    sorted(raw["subproject_name"].dropna().astype(str).str.strip().unique().tolist())
    if "subproject_name" in raw.columns else []
)

# Detect which subprojects are missing an ID from the auto-lookup
_missing_ids = [n for n in _unique_subprojects if n not in _auto_lookup or pd.isna(_auto_lookup.get(n))]

if _auto_lookup and not _missing_ids:
    sub_df = (
        pd.DataFrame([{"subproject_name": k, "subproject_id": v} for k, v in _auto_lookup.items()])
        .sort_values("subproject_name").reset_index(drop=True)
    )
    st.success(f"Subproject IDs auto-detected from RAW for all {len(sub_df)} subproject(s).")
    st.dataframe(sub_df, use_container_width=True, hide_index=True)
    _manual_override = False
else:
    if _auto_lookup:
        st.warning(f"Subproject IDs partially detected. Missing IDs for: **{', '.join(_missing_ids)}**. Fill in below.")
    else:
        st.info("No `subproject_id` column found in RAW. Enter subproject IDs manually.")
    _manual_override = True

# Always show override expander when there are missing IDs; collapsed otherwise
with st.expander("✏️ Set / override subproject IDs", expanded=_manual_override):
    st.caption("Values here override anything detected from RAW.")
    _oc = st.columns(min(len(_unique_subprojects), 3) or 1)
    _edited_lookup = dict(_auto_lookup)  # start from auto-detected
    for _i, _name in enumerate(_unique_subprojects):
        _default = _auto_lookup.get(_name, "")
        _val = _oc[_i % 3].text_input(
            label=_name,
            value=str(int(_default)) if _default != "" and not pd.isna(_default) else "",
            placeholder="Enter numeric ID",
            key=f"spid_{_name}",
        )
        if _val.strip():
            try:
                _edited_lookup[_name] = int(_val.strip())
            except ValueError:
                _edited_lookup[_name] = _val.strip()

subproject_lookup = _edited_lookup

# =============================================================================
# PROJECT NAME
# =============================================================================

st.subheader("Project Name")

_detected_project_name = detect_project_name(raw)

if _detected_project_name:
    st.success(f"Project name detected from RAW: **{_detected_project_name}**")
    project_name = _detected_project_name
else:
    st.info("No `project_name` column found in RAW. Enter the project name manually.")
    project_name = st.text_input(
        "Project Name",
        value=st.session_state.get("project_name_input", ""),
        placeholder="e.g. Nespresso - Indonesia Aceh",
        key="project_name_input",
    )
    if not project_name.strip():
        st.warning("Project name is empty — it will be blank in the output.")

# =============================================================================
# TRACK UPLOAD
# =============================================================================

st.subheader("Parcel GPS Track")

track_upload = st.file_uploader(
    "Upload parcel GPS track (.xlsx or .csv)",
    type=["xlsx", "csv"]
)

track_df = None

if track_upload:
    try:
        if track_upload.name.lower().endswith(".csv"):
            track_df = pd.read_csv(track_upload)
        else:
            track_df = pd.read_excel(track_upload)
        track_df = normalize_columns(track_df)
        if "geometry" not in track_df.columns:
            st.error("Missing geometry column")
            track_df = None
        else:
            st.success(f"{len(track_df):,} parcel tracks loaded")
    except Exception as e:
        st.error(f"Track upload failed: {e}")

# =============================================================================
# MAP  — always shown, zooms to fit all data
# =============================================================================

st.subheader("GPS Map")

point_df   = build_point_layer(raw)
polygon_df = build_polygon_layer(track_df)

layers = []

if not point_df.empty:
    layers.append(pdk.Layer(
        "ScatterplotLayer",
        data=point_df,
        get_position="[lon, lat]",
        get_radius=1,
        radius_min_pixels=2,
        radius_max_pixels=6,
        opacity=0.85,
        pickable=True,
        filled=True,
        stroked=False,
        get_fill_color=[255, 80, 0],
    ))

if not polygon_df.empty:
    layers.append(pdk.Layer(
        "PolygonLayer",
        data=polygon_df,
        get_polygon="polygon",
        filled=True,
        stroked=True,
        wireframe=True,
        get_fill_color=[0, 120, 255, 30],
        get_line_color=[0, 100, 255],
        get_line_width=3,
        line_width_min_pixels=1,
        pickable=True,
    ))

# Compute bounding box so the map auto-fits all data
if not point_df.empty:
    min_lat = point_df["lat"].min()
    max_lat = point_df["lat"].max()
    min_lon = point_df["lon"].min()
    max_lon = point_df["lon"].max()
    center_lat = (min_lat + max_lat) / 2
    center_lon = (min_lon + max_lon) / 2

    # estimate zoom from span
    lat_span = max_lat - min_lat
    lon_span = max_lon - min_lon
    span = max(lat_span, lon_span)
    if span < 0.01:
        zoom = 14
    elif span < 0.05:
        zoom = 12
    elif span < 0.2:
        zoom = 10
    elif span < 1:
        zoom = 8
    else:
        zoom = 6
else:
    center_lat, center_lon, zoom = 0, 0, 3

view_state = pdk.ViewState(
    latitude=center_lat,
    longitude=center_lon,
    zoom=zoom,
    pitch=0,
)

deck = pdk.Deck(
    layers=layers,
    initial_view_state=view_state,
    map_style="https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
    tooltip={"text": "{label}"},
)

st.pydeck_chart(deck, use_container_width=True)
st.caption("🔴 Red dots = RAW subplot GPS points  |  🔵 Blue polygons = Parcel GPS tracks (if uploaded)")

# GIS export — only build files when the user clicks the button
if st.button("⬇️ Generate GIS Files", key="gen_gis"):
    combined_geojson = build_combined_geojson(raw, track_df)
    st.session_state["combined_geojson"] = combined_geojson
    st.session_state["combined_kml"]     = geojson_to_kml(combined_geojson)
    if GIS_AVAILABLE:
        shp_zip = export_shapefile(raw, track_df)
        st.session_state["shp_zip"] = shp_zip.read() if shp_zip else None
    st.success("GIS files ready — click a button below to download.")

if "combined_geojson" in st.session_state:
    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button("Download GeoJSON", st.session_state["combined_geojson"],
                           file_name="combined_map.geojson", mime="application/geo+json")
    with col2:
        st.download_button("Download KML", st.session_state["combined_kml"],
                           file_name="combined_map.kml", mime="application/vnd.google-earth.kml+xml")
    with col3:
        if GIS_AVAILABLE and st.session_state.get("shp_zip"):
            st.download_button("Download SHP ZIP", st.session_state["shp_zip"],
                               file_name="combined_shapefile.zip", mime="application/zip")
        elif not GIS_AVAILABLE:
            st.caption("Install geopandas for SHP export")

# =============================================================================
# TRANSFORMATION TOGGLES
# =============================================================================

st.subheader("Transformation Options")

generate_farmer        = st.toggle("Generate farmer_name",               value=True)
generate_parcel        = st.toggle("Generate parcel",                    value=True)
generate_pulldata      = st.toggle("Generate pulldata_sampling_biomass", value=True)
generate_subplot       = st.toggle("Generate subplot_gpspoint",          value=True)
generate_subplot_parcel = st.toggle("Generate subplot_and_parcel (GPS points + parcel tracks combined)", value=False)

# =============================================================================
# RUN TRANSFORMATION
# =============================================================================

if st.button("Run Transformation", type="primary"):
    results = {}
    if generate_farmer:
        results["farmer_name"] = build_farmer_name(raw, subproject_lookup)
    if generate_parcel:
        results["parcel"] = build_parcel(raw)
    if generate_pulldata:
        results["pulldata_sampling_biomass"] = build_pulldata(raw, subproject_lookup, project_name)
    if generate_subplot:
        results["subplot_gpspoint"] = build_subplot_gpspoint(raw)
    if generate_subplot_parcel:
        results["subplot_and_parcel"] = build_subplot_and_parcel(raw, track_df)
    st.session_state["results"] = results
    st.success("Transformation complete")

# =============================================================================
# RESULTS
# =============================================================================

if "results" in st.session_state:
    results = st.session_state["results"]

    st.subheader("Results")

    # Row count summary
    summary_cols = st.columns(len(results))
    for col, (name, df) in zip(summary_cols, results.items()):
        col.metric(name, f"{len(df):,} rows")

    st.markdown("")

    tabs = st.tabs(list(results.keys()))
    for tab, (name, df) in zip(tabs, results.items()):
        with tab:
            st.dataframe(df, use_container_width=True, hide_index=True)

    # =========================================================================
    # MISSING VALUE RECAP
    # =========================================================================

    st.subheader("Missing Value Recap")

    whole_empty, partial_empty = check_empty_columns(results, raw)

    if whole_empty:
        st.warning("Whole column empty (no data at all)")
        st.dataframe(pd.DataFrame(whole_empty), use_container_width=True, hide_index=True)
    else:
        st.success("No fully empty columns")

    if partial_empty:
        st.warning("Partially missing columns")
        st.dataframe(pd.DataFrame(partial_empty), use_container_width=True, hide_index=True)
    else:
        st.success("No partially missing columns")

    # =========================================================================
    # DOWNLOAD
    # =========================================================================

    st.subheader("Download")

    # ── Option A: single Excel with all tabs ─────────────────────────────────
    st.markdown("**Excel (all tabs in one file)**")
    excel_bytes = to_excel(results)
    st.download_button(
        "⬇️ Download Excel",
        excel_bytes,
        file_name="pulldata_transformed.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")

    # ── Option B: individual CSV per tab ─────────────────────────────────────
    # Each button uses a fixed filename so the browser always overwrites the
    # previous download rather than saving as "parcel (2).csv" etc.
    st.markdown("**CSV (one file per tab)**")
    csv_cols = st.columns(len(results))
    for col, (tab_name, df) in zip(csv_cols, results.items()):
        df = enforce_label_fallback(df)
        csv_bytes = df.to_csv(index=False).encode("utf-8")
        col.download_button(
            label=f"⬇️ {tab_name}.csv",
            data=csv_bytes,
            file_name=f"{tab_name}.csv",   # fixed name → browser overwrites
            mime="text/csv",
            use_container_width=True,
            key=f"csv_{tab_name}",
        )